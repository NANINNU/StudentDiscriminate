#Library Import
import os, sys
import time
import newResource_rc
from PyQt5 import *
from PyQt5 import uic
from PyQt5.QtWidgets import *
from PyQt5.QtCore import *
from PyQt5.QtGui import *
import openpyxl as op


#UI Packaging
def resourcePath(relativePath):
    basePath = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(basePath, relativePath)


#UI파일 연결
#단, UI파일은 Python 코드 파일과 같은 디렉토리에 위치해야한다.
form = resourcePath("studentDiscriminate 1.3ver.ui")
form_class = uic.loadUiType(form)[0]

#CONSTANT LOAD DEFINE
_GLOBAL_LOAD_EXCEL_PAID = 'PAID'
_GLOBAL_LOAD_EXCEL_SURVEY = 'SURVEY'
_GLOBAL_LOAD_EXCEL_STUDENT = 'STUDENT'

#CONSTANT WRITE DEFINE
_GLOBAL_WRITE_EXCEL_PAID = 'PAID'
_GLOBAL_WRITE_EXCEL_UNPAID = 'UNPAID'


########################################################################################
#화면을 띄우는데 사용되는 Class 선언
class MainWindow(QDialog, form_class) :
    def __init__(self) :
        super().__init__()
        self.setupUi(self)
        self.initUI()   
        
    # ##################################################################################
    #UI Initialize
    def initUI(self):
        #Dialog Size Fixed
        #self.setFixedSize(QSize(1063, 584))
        #Setup Exit button color
        self.programExit.setStyleSheet("color: red;"
                             "border-style: solid;"
                             "border-width: 2px;"
                             "border-color: #FA8072;"
                             "border-radius: 3px")
        
        #setFont
        QFontDatabase.addApplicationFont(os.path.join(r'E:\Project\studentDiscriminate\src\NanumBarunpenB.otf'))
        
        font = QFont('NanumBarunpenB.otf', 10)
        QApplication.setFont(font)
        
        #setWindowIcon
        window_ico = resourcePath('appIcon.ico')
        self.setWindowIcon(QIcon(window_ico))

        #Template Button Tooltip
        self.studentTemplate.setToolTip("Template 다운로드하여 재학생 명단을 작성합니다.")
        self.paidTemplate.setToolTip('Template 다운로드하여 납부자 명단을 작성합니다.')
        self.surveyTemplate.setToolTip('Template 다운로드하여 설문자 명단을 작성합니다.')
    
        # self.studentPBar.hide()
        # self.paidPBar_2.hide()
        # self.surveyPBar.hide()
        
        
        #재학생 GroupBox
        #명단 불러오기 버튼을 눌렀을때 selectFunctionA 매서드 동작
        self.loadstudentList.clicked.connect(self.selectFunctionA)
        #File Template Download 버튼을 눌렀을 때 studentTemplateFunc 매서드 동작
        self.studentTemplate.clicked.connect(self.studentTempleteFunc)
        
        
        #납부자 GroupBox
        #명단 불러오기 버튼을 눌렀을때 selectFunctionP 매서드 동작
        self.loadpaidList.clicked.connect(self.selectFunctionP)
        #File Template Download 버튼을 눌렀을 때 paidTemplateFunc 매서드 동작
        self.paidTemplate.clicked.connect(self.paidTemplateFunc)
        
        
        #설문자 GroupBox
        #명단 불러오기 버튼을 눌렀을때 selectFunctionS 매서드 동작
        self.loadsurveyList.clicked.connect(self.selectFunctionS)
        #File Template Download 버튼을 눌렀을 때 surveyTemplateFunc 매서드 동작
        self.surveyTemplate.clicked.connect(self.surveyTemplateFunc)
        
        #대조하기 버튼을 눌렀을때 Compare 매서드 동작
        self.compareList.clicked.connect(self.Compare)
        
        #명단 추출하기 버튼을 눌렀을 때 writeExcel 매서드 동작
        self.printList.clicked.connect(self.writeExcel)
        

        
        #명단 초기화 버튼을 눌렀을 때 reset 매서드 동작
        self.resetList.clicked.connect(self.reset)
        
        #종료 버튼을 눌렀을 때 programExit 매서드 동작
        self.programExit.clicked.connect(self.exit)
        

        #재학생 명단 리스트화
        self.studentList = []

        #납부자 명단 리스트화
        self.paidList = []
        
        #설문자 명단 리스트화
        self.surveyList = []  
        self.CompareYN = False
        
        self.show()
    
    #Load All Student List(.xlsx File)
    def selectFunctionA(self):
        self.loadExcel(mode = _GLOBAL_LOAD_EXCEL_STUDENT)
    #Download All Student List
    def studentTempleteFunc(self):
        self.createTemplate(mode = _GLOBAL_LOAD_EXCEL_STUDENT) 
    
    #Load PAID Student List(.xlsx FILE)
    def selectFunctionP(self):
        self.loadExcel(mode = _GLOBAL_LOAD_EXCEL_PAID)
    #DOWNLOAD PAID STUDENT LIST TEMPLATE FILE
    def paidTemplateFunc(self):
        self.createTemplate(mode = _GLOBAL_LOAD_EXCEL_PAID)
    
    #LOAD SURVEY STUDENT LIST(.xlsx FILE)
    def selectFunctionS(self):
        self.loadExcel(mode = _GLOBAL_LOAD_EXCEL_SURVEY)
    #DOWNLOAD SURVEY STUDENT LIST TEMPLATE FILE
    def surveyTemplateFunc(self):
        self.createTemplate(mode = _GLOBAL_LOAD_EXCEL_SURVEY)
    
    #########################################################################################################
    #MAKING TEMPLATE FILE MODULE
    def createTemplate(self, mode):
        fileName = ''
        
        try:
            wb = op.Workbook()
            sheet = wb.active
            
            # 납부자/미납자 분리
            if mode == _GLOBAL_WRITE_EXCEL_PAID:
                fileName = '납부자명단_Template.xlsx'
                sheet.append(["이름","학번"])
            elif mode == _GLOBAL_LOAD_EXCEL_SURVEY:
                fileName = '설문자명단_Template.xlsx'
                sheet.append(["이름","학번","납부여부"])
            elif mode == _GLOBAL_LOAD_EXCEL_STUDENT:
                fileName = '재학생 명단_Templete.xlsx'
                sheet.append(["이름", "학번"])

            sheet.append(["홍길동","12345678"])
            wb.save(fileName)
            
            #Message: Complete writing excel file
            QMessageBox.information(self, "알림", f"Template이 생성되었습니다.({fileName})")
            
            #Template File 화면 표시
            self.txtOutputFileName.setPlainText(os.getcwd() + fileName)
        
        except Exception as e:
            QMessageBox.critical(self, "오류", f"Write Excel Error: {e}")
            print(f'Exception error occurred: {e}')


    ''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
    #LOAD AND PRINT PAID/SURVEY STUDENT LIST AT QTABLE WIDGET MODULE
    def loadExcel(self, mode):
        try:
            self.CompareYN = False
            path = QFileDialog.getOpenFileName(self, 'Open File', '', 'xlsx File(*.xlsx)')[0]

            if path == '':
                QMessageBox.warning(self, "경고", "파일경로를 읽어올 수 없습니다")
                return
            
            # 파일경로/파일명 분리
            dirPath, Name = os.path.split(path)

            #위 절대 경로 활용해 openpyxl workbook 객체 생성
            wb = op.load_workbook(path)
            sheet = wb.active
            
            #excel 파일의 행과 열 입력
            maxCol = sheet.max_column
            maxRow = sheet.max_row
            sheetData = list(sheet.values)          
            
            if mode == _GLOBAL_LOAD_EXCEL_PAID:
                
                #PRINT FILE PATH
                self.paidAddress.setPlainText(dirPath)
                
                #PRINT FILE NAME
                self.paidfileName.setText(Name)
                
                self.paidTable.setRowCount(maxRow-1)
                self.paidTable.setColumnCount(maxCol)
                
                #각 행의 제목 입력
                self.paidTable.setHorizontalHeaderLabels(sheetData[0])
                title = list(sheetData[0])
                if title[0] != "이름" or title[1] != "학번":
                    QMessageBox.information(self, "알림", "명단의 첫번째 행에는 '이름','학번'이 있어야 합니다.")
                    return
                
                #LOADING EXCEL FILE PROGRESS BAR
                row_index = 0
                self.paidPBar.show()
                self.paidPBar.maximum = maxRow
                self.paidPBar.setValue(0)
                
                #엑셀파일 추출
                for value_tuple in sheetData[1:]:
                    self.paidList.append(list(value_tuple))#list에 tuple로 저장되는 문제가 있어서 tuple을 list로 변환
                    
                    time.sleep(0.001)
                    self.paidPBar.setValue(int(((row_index+1)/(maxRow-1))*100))
                    
                    col_index = 0
                    for value in value_tuple:
                        self.paidTable.setItem(row_index , col_index, QTableWidgetItem(str(value)))
                        col_index += 1
                    row_index += 1
                self.paidPBar.hide()
                
            elif mode == _GLOBAL_LOAD_EXCEL_SURVEY:
                self.surveyAddress.setPlainText(dirPath)
                self.surveyfileName.setText(Name)
                self.surveyTable.setRowCount(maxRow-1)
                self.surveyTable.setColumnCount(maxCol)

                #각 행의 제목 입력
                self.surveyTable.setHorizontalHeaderLabels(sheetData[0])
                title = list(sheetData[0])
                if title[0] != "이름" or title[1] != "학번" or title[2] != "납부여부" :
                    QMessageBox.information(self, "알림", "명단의 첫번째 행에는 '이름','학번','납부여부'가 있어야 합니다.")
                    return
                
                #엑셀파일 추출
                row_index = 0
                
                #MAKEING EXCEL FILE PROGRESS BAR                
                self.surveyPBar.show()
                self.surveyPBar.maximum = maxRow
                self.surveyPBar.setValue(0)
                
                for value_tuple in sheetData[1:]:
                    self.surveyList.append(list(value_tuple))#list에 tuple로 저장되는 문제가 있어서 tuple을 list로 변환
                    
                    time.sleep(0.001)
                    self.surveyPBar.setValue(int(((row_index+1)/(maxRow-1))*100))
                    
                    col_index = 0
                    for value in value_tuple:
                        if value==None:
                            value=''
                        self.surveyTable.setItem(row_index , col_index, QTableWidgetItem(str(value)))
                        col_index += 1
                    row_index += 1
                self.surveyPBar.hide()
            
            elif mode == _GLOBAL_LOAD_EXCEL_STUDENT:
                self.studentAddress.setPlainText(dirPath)
                self.studentfileName.setText(Name)
                self.studentTable.setRowCount(maxRow-1)
                self.studentTable.setColumnCount(maxCol)

                #각 행의 제목 입력
                self.studentTable.setHorizontalHeaderLabels(sheetData[0])
                title = list(sheetData[0])
                if title[0] != "이름" or title[1] != "학번":
                    QMessageBox.information(self, "알림", "명단의 첫번째 행에는 '이름','학번'이 있어야 합니다.")
                    return
                
                #엑셀파일 추출
                row_index = 0
                
                #MAKEING EXCEL FILE PROGRESS BAR                
                self.studentPBar.show()
                self.studentPBar.maximum = maxRow
                self.studentPBar.setValue(0)
                
                for value_tuple in sheetData[1:]:
                    self.studentList.append(list(value_tuple))#list에 tuple로 저장되는 문제가 있어서 tuple을 list로 변환
                    
                    time.sleep(0.001)
                    self.studentPBar.setValue(int(((row_index+1)/(maxRow-1))*100))
                    
                    col_index = 0
                    for value in value_tuple:
                        if value==None:
                            value=''
                        self.studentTable.setItem(row_index , col_index, QTableWidgetItem(str(value)))
                        col_index += 1
                    row_index += 1
                self.studentPBar.hide()
                
        except Exception as e:
            QMessageBox.critical(self, "오류", f"File load fail: {e}")
            print(f'Exception error occurred: {e}')
    ''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
#########################################################################################################################

    #COMPARE PAID/UNPAID STUDENT
    def Compare(self):
        
        if self.paidList == [] or self.surveyList == []:
            QMessageBox.information(self, "알림", "명단을 먼저 불러오십시오")
            return
        
        try:        
            for row, val in enumerate(self.surveyList):
                self.surveyList[row][2] = self.checkStudent(val[1])
             
            
            row_index=0      
            for rowData in self.surveyList:
                col_index = 0
                for cell in rowData:
                    self.surveyTable.setItem(row_index , col_index, QTableWidgetItem(str(cell)))
                    col_index += 1
                row_index += 1

        
            self.CompareYN = True
            QMessageBox.information(self, "알림", "대조가 완료되었습니다.")
            self.txtOutputFileName.clear()
            
        except Exception as e:
            QMessageBox.critical(self, "오류", f"Compare Error: {e}")
            print(f'Exception error occurred: {e}')
##################################################################################################################
    
    def checkStudent(self, surveyed):
        for student in self.studentList:
            if surveyed == student[1]:
                return self.checkPaid(surveyed)
        return '휴학생'
            
###################################################################################################################
    def checkPaid(self, student):
        for paid in self.paidList:
            if student == paid[1]:
                return '납부자'
        return '미납부자'
####################################################################################################################
    

    #WRITE AND SAVE EXCEL FILE
    def writeExcel(self, mode):
        if self.paidList == [] or self.surveyList == []:
            QMessageBox.information(self, "알림", "명단을 먼저 불러오십시오", QMessageBox.Ok)
            return

        if self.CompareYN == False:
            QMessageBox.information(self, "알림", "명단을 먼저 대조하십시오", QMessageBox.Ok)
            return
        fileName = 'OO행사 참여자 명단.xlsx'
        
        try:
            
            wb = op.Workbook()
            sheet = wb.active
            sheet.append(["이름","학번","납부여부"])
            
            for Rows in range(len(self.surveyList)):
                sheet.append([self.surveyList[Rows][0], self.surveyList[Rows][1], self.surveyList[Rows][2]])

            wb.save(fileName)
            
            #Message: Complete writing excel file
            QMessageBox.information(self, "알림", "파일이 내보내졌습니다")
            
            #내보내기 파일 화면 표시
            self.txtOutputFileName.setPlainText(os.getcwd() + fileName)
            
        
        except Exception as e:
            QMessageBox.critical(self, "오류", f"Write Excel Error: {e}")
            print(f'Exception error occurred: {e}')
            
    #Reset List
    def reset(self):
        resetQ = QMessageBox.question(self, "주의", "명단을 모두 초기화 합니다.\n명단을 초기화하시겠습니까?", QMessageBox.Yes|QMessageBox.No)
        
        if resetQ == QMessageBox.Yes:
            self.surveyList = []
            self.paidList = []
            self.paidTable.clearContents()
            self.paidTable.setRowCount(0)
            self.paidTable.setColumnCount(0)
            self.surveyTable.clearContents()
            self.surveyTable.setRowCount(0)
            self.surveyTable.setColumnCount(0)
            QMessageBox.information(self, "알림", "초기화되었습니다.\n명단을 다시 불러오십시오.")
    
    
    #EXIT PROGRAM
    def exit(self):
        re = QMessageBox.question(self, "확인", "프로그램을 종료 하시겠습니까?", QMessageBox.Yes|QMessageBox.No)

        if re == QMessageBox.Yes:
            QApplication.quit()
               
# ######################################################################################
if __name__ == "__main__" :
    #QApplication : 프로그램을 실행시켜주는 클래스
    app = QApplication(sys.argv) 

    #MainWindow 인스턴스 생성
    myWindow = MainWindow() 

    #프로그램 화면을 보여주는 코드
    myWindow.show()

    #프로그램을 이벤트루프로 진입시키는(프로그램을 작동시키는) 코드
    sys.exit(app.exec_())